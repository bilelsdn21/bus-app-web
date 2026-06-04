"""
Seed the DB from the existing webapp data/*.json files.
Run once:  python -m app.seed
Buses are de-duplicated by PLATE so historical name variants collapse into one
vehicle (the bug that lost 1,440 TND in the Excel app can't happen here).
"""
import json
import re
from pathlib import Path
from datetime import datetime, date

from .core.database import Base, engine, SessionLocal
from . import models

# data/ lives in the original project root: BUS PROJECT/data
PROJECT_ROOT = Path(__file__).resolve().parents[3]
DATA_DIR = PROJECT_ROOT / "data"

# Excel file to pull destination categories + prices from (Parametres / tblParams)
DEST_SOURCE_XLSX = PROJECT_ROOT / "final.xlsm"


def seed_buses_from_excel(db):
    """
    PARAMS come from final.xlsm -> Parametres -> tblBuses (the hand-tuned source of truth).
    Returns {plate: Bus}. The user can later edit buses in the app.
    """
    out = {}
    try:
        import openpyxl
        from openpyxl.utils import range_boundaries
    except ImportError:
        print("  (openpyxl missing — cannot read bus params)")
        return out
    if not DEST_SOURCE_XLSX.exists():
        print(f"  (no {DEST_SOURCE_XLSX.name} — cannot read bus params)")
        return out
    wb = openpyxl.load_workbook(str(DEST_SOURCE_XLSX), data_only=True)
    ws = wb["Parametres"]
    table = None
    for t in ws.tables.values():
        c0, r0, c1, r1 = range_boundaries(t.ref)
        head = [str(ws.cell(row=r0, column=c).value or "").strip() for c in range(c0, c1 + 1)]
        if "Nom du Vehicule" in head:
            table = (c0, r0, c1, r1, head); break
    if not table:
        wb.close(); return out
    c0, r0, c1, r1, head = table
    def col(name):
        return c0 + head.index(name) if name in head else None
    ni, ti = col("Nom du Vehicule"), col("Type")
    li, ri, di = col("Loyer Mensuel (TND)"), col("Region"), col("Distance (km)")
    for r in range(r0 + 1, r1 + 1):
        name = ws.cell(row=r, column=ni).value if ni else None
        if not name:
            continue
        name = str(name).strip()
        def num(idx):
            v = ws.cell(row=r, column=idx).value if idx else 0
            try: return float(v or 0)
            except (TypeError, ValueError): return 0.0
        bus = models.Bus(
            name=name, plate=plate_of(name),
            type=str(ws.cell(row=r, column=ti).value or "MICRO").strip() if ti else "MICRO",
            region=str(ws.cell(row=r, column=ri).value or "").strip() if ri else "",
            loyer=num(li), distance=num(di),
        )
        db.add(bus); db.flush()
        out[plate_of(name) or name] = bus
    wb.close()
    return out


def seed_destinations(db):
    """Import Categorie | Destination | Prix MICRO/OTOKAR/BUS from the Excel tblParams."""
    try:
        import openpyxl
        from openpyxl.utils import range_boundaries
    except ImportError:
        print("  (openpyxl missing — skipping destination import)")
        return 0
    if not DEST_SOURCE_XLSX.exists():
        print(f"  (no {DEST_SOURCE_XLSX.name} — skipping destination import)")
        return 0
    wb = openpyxl.load_workbook(str(DEST_SOURCE_XLSX), data_only=True)
    if "Parametres" not in wb.sheetnames:
        wb.close(); return 0
    ws = wb["Parametres"]
    table = None
    for t in ws.tables.values():
        c0, r0, c1, r1 = range_boundaries(t.ref)
        head = [ws.cell(row=r0, column=c).value for c in range(c0, c1 + 1)]
        if "Destination" in [str(h).strip() if h else "" for h in head]:
            table = (c0, r0, c1, r1, head); break
    if not table:
        wb.close(); return 0
    c0, r0, c1, r1, head = table
    hl = [str(h).strip() if h else "" for h in head]
    def col(name):
        return c0 + hl.index(name) if name in hl else None
    ci, di = col("Categorie"), col("Destination")
    mi, oi, bi = col("Prix MICRO"), col("Prix OTOKAR"), col("Prix BUS")
    n = 0
    for r in range(r0 + 1, r1 + 1):
        dest = ws.cell(row=r, column=di).value if di else None
        if not dest:
            continue
        def num(idx):
            v = ws.cell(row=r, column=idx).value if idx else 0
            try: return float(v or 0)
            except (TypeError, ValueError): return 0.0
        db.add(models.Destination(
            category=str(ws.cell(row=r, column=ci).value or "").strip() if ci else "",
            name=str(dest).strip(),
            price_micro=num(mi), price_otokar=num(oi), price_bus=num(bi),
        ))
        n += 1
    wb.close()
    return n


def plate_of(name: str) -> str:
    m = re.search(r"\d+\s*TU\s*\d+", name or "")
    return re.sub(r"\s+", " ", m.group(0)).strip() if m else ""


def type_order(t: str) -> int:
    return {"BUS": 1, "OTOKAR": 2, "MICRO": 3}.get((t or "").upper(), 4)


def region_order(r: str) -> int:
    return {"SOUSSE": 1, "DJERBA": 2}.get((r or "").upper(), 3)


def run():
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    db = SessionLocal()

    if not db.query(models.Config).first():
        db.add(models.Config(cut_morn=13, cut_night=22))

    files = sorted(DATA_DIR.glob("*.json"))
    print(f"Found {len(files)} month file(s) in {DATA_DIR}")

    cut_morn, cut_night = 13, 22
    for f in files:
        d = json.loads(f.read_text(encoding="utf-8"))
        cut_morn = d.get("cut_morn", cut_morn)
        cut_night = d.get("cut_night", cut_night)

    # PARAMS (buses) come from final.xlsm — the authoritative, hand-tuned source.
    bus_by_plate = seed_buses_from_excel(db)
    print(f"  Loaded {len(bus_by_plate)} buses from {DEST_SOURCE_XLSX.name}")
    db.flush()

    # sort buses: region then type then name
    for i, bus in enumerate(sorted(bus_by_plate.values(),
                                   key=lambda x: (region_order(x.region), type_order(x.type), x.name))):
        bus.sort_order = i

    # bookings (match to bus by plate; create a minimal bus if a plate is unknown,
    # so historical revenue is never silently dropped)
    n_book, n_extra = 0, 0
    for f in files:
        data = json.loads(f.read_text(encoding="utf-8"))
        for bk in data.get("bookings", []):
            ds = bk.get("date")
            if not ds:
                continue
            try:
                d = datetime.strptime(ds, "%Y-%m-%d").date()
            except ValueError:
                continue
            raw_name = (bk.get("bus") or "").strip()
            pl = plate_of(raw_name)
            bus = bus_by_plate.get(pl) or bus_by_plate.get(raw_name)
            if not bus and raw_name:
                bus = models.Bus(name=raw_name, plate=pl, type="MICRO", region="", loyer=0, distance=0)
                db.add(bus); db.flush()
                bus_by_plate[pl or raw_name] = bus
                n_extra += 1
            if not bus:
                continue
            def fnum(x):
                try: return float(x)
                except (TypeError, ValueError): return 0.0
            def inum(x):
                try: return int(float(x))
                except (TypeError, ValueError): return 0
            db.add(models.Booking(
                date=d, bus_id=bus.id, type=bk.get("type", "Booking"),
                destination=bk.get("destination", ""), client=bk.get("client", ""),
                pax=inum(bk.get("pax")), unit_price=fnum(bk.get("price")),
                total=fnum(bk.get("total")), notes=bk.get("notes", ""),
                heure_debut=bk.get("heure_debut", "") or bk.get("heure", ""),
                heure_fin=bk.get("heure_fin", ""),
            ))
            n_book += 1

    n_dest = seed_destinations(db)

    # default contract + fuel per bus per month that has data (user edits these to real dates)
    import calendar as _cal
    db.flush()
    seen = set()  # (bus_id, year, month)
    n_contract = n_fuel = 0
    for bk in db.query(models.Booking).all():
        key = (bk.bus_id, bk.date.year, bk.date.month)
        if key in seen:
            continue
        seen.add(key)
        bus = db.get(models.Bus, bk.bus_id)
        y, m = bk.date.year, bk.date.month
        last = _cal.monthrange(y, m)[1]
        db.add(models.Contract(bus_id=bk.bus_id, start_date=date(y, m, 1),
                               end_date=date(y, m, last), loyer=bus.loyer,
                               label=f"{m:02d}/{y}"))
        db.add(models.FuelMonth(bus_id=bk.bus_id, year=y, month=m, estimated=0.0, actual=None))
        n_contract += 1; n_fuel += 1

    cfg = db.query(models.Config).first()
    cfg.cut_morn, cfg.cut_night = cut_morn, cut_night
    db.commit()
    print(f"Seeded {len(bus_by_plate)} buses ({n_extra} extra from bookings), {n_book} bookings, "
          f"{n_dest} destinations, {n_contract} contracts, {n_fuel} fuel-months. "
          f"Cutoffs {cut_morn}/{cut_night}.")
    db.close()


if __name__ == "__main__":
    run()
